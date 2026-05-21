"""
Isolates classification logic and boundary detection.
Adheres to the Open-Closed Principle (OCP).
"""
import logging
import re
from typing import Dict, List, Set, Tuple, Any
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

from config import PANEL_ROW, HEADER_ROW, PANEL_START_COL, END_MARKER

def _is_empty(val: Any) -> bool:
    """Evaluates if a cell value constitutes an empty or null equivalent state."""
    if val is None:
        return True
    return str(val).strip().upper() in ("", "0", "0.0", "NONE", "NULL", "-")

def _is_numeric(val: Any) -> bool:
    """Evaluates if a cell value is strictly numerical."""
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return True
    try:
        float(val)
        return True
    except ValueError:
        return False

def _sanitize_sheet_title(title: str) -> str:
    """
    Strips Excel-forbidden characters (\\, /, *, ?, [, ]) and enforces the 31-character limit.
    """
    safe_title = re.sub(r'[\\/*?:\[\]]', '_', title)
    return safe_title[:31]

def detect_columns(ws: Worksheet, logger: logging.Logger) -> Tuple[Dict[str, List[int]], List[int], Dict[str, int]]:
    """
    Scans the worksheet to map panel boundaries, series prefixes, and static headers.
    """
    panel_cols: Dict[str, List[int]] = {}
    all_panel_cols: List[int] = []
    header_map: Dict[str, int] = {}

    end_col: int = -1
    for col in range(PANEL_START_COL, ws.max_column + 1):
        val1 = str(ws.cell(row=1, column=col).value or "").strip().upper()
        val2 = str(ws.cell(row=PANEL_ROW, column=col).value or "").strip().upper()
        if END_MARKER in val1 or END_MARKER in val2:
            end_col = col
            break

    if end_col == -1:
        logger.warning(f"Boundary marker '{END_MARKER}' not found. Scanning to max boundary.")
        end_col = ws.max_column + 1

    for col in range(PANEL_START_COL, end_col):
        raw_val = ws.cell(row=PANEL_ROW, column=col).value
        val = str(raw_val or "").strip()
        if val:
            all_panel_cols.append(col)
            prefix = val.split("-")[0].strip()
            if prefix not in panel_cols:
                panel_cols[prefix] = []
            panel_cols[prefix].append(col)

    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=HEADER_ROW, column=col).value
        if val:
            header_map[str(val).strip().upper()] = col

    return panel_cols, all_panel_cols, header_map

def classify_panels(wb: Workbook, ws: Worksheet, panel_cols: Dict[str, List[int]],
                   all_panel_cols: List[int], header_map: Dict[str, int], logger: logging.Logger) -> int:
    """
    Duplicates the source worksheet per series, systematically pruning unrelated rows
    and columns while preserving format, color integrity, and static components.
    """
    sheets_created: int = 0
    sno_col: int = header_map.get('SNO', 1)
    desc_col: int = header_map.get('DESCRIPTION', 2)
    cat_col: int = header_map.get('CAT NO.', 6)

    for prefix, prefix_panel_cols in panel_cols.items():
        logger.info(f"Generating isolated structure for series: {prefix}")

        ws_new: Worksheet = wb.copy_worksheet(ws)
        ws_new.title = _sanitize_sheet_title(prefix)

        data_rows_to_delete: Set[int] = set()

        for r in range(HEADER_ROW + 1, ws.max_row + 1):
            sno_val = ws.cell(row=r, column=sno_col).value
            cat_val = ws.cell(row=r, column=cat_col).value

            is_data_row = _is_numeric(sno_val) or not _is_empty(cat_val)
            if is_data_row:
                has_prefix_qty = any(
                    not _is_empty(ws.cell(row=r, column=c).value) and ws.cell(row=r, column=c).value != 0
                    for c in prefix_panel_cols
                )
                if not has_prefix_qty:
                    data_rows_to_delete.add(r)

        rows_to_delete: List[int] = sorted(list(data_rows_to_delete), reverse=True)
        for r in rows_to_delete:
            ws_new.delete_rows(r)

        cols_to_delete: List[int] = sorted([c for c in all_panel_cols if c not in prefix_panel_cols], reverse=True)
        for c in cols_to_delete:
            ws_new.delete_cols(c)

        for c in range(1, ws_new.max_column + 1):
            val1 = str(ws_new.cell(row=1, column=c).value or "").strip().upper()
            val2 = str(ws_new.cell(row=PANEL_ROW, column=c).value or "").strip().upper()
            if END_MARKER in val1 or END_MARKER in val2:
                for mr in list(ws_new.merged_cells.ranges):
                    if mr.min_col == c and (mr.min_row == 1 or mr.min_row == PANEL_ROW):
                        ws_new.merged_cells.remove(mr)

                ws_new.merge_cells(start_row=PANEL_ROW, start_column=c, end_row=PANEL_ROW, end_column=c+1)
                break

        sheets_created += 1

    return sheets_created