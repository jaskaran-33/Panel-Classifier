"""
Handles workbook ingestion, applies resource limits via bounds checking,
and routes structural data to the classifier.
"""
from pathlib import Path
import logging
import openpyxl

from config import OUTPUT_DIR, OUTPUT_FILE, MAX_ROWS, MAX_COLS
from utils.classifier import detect_columns, classify_panels
from utils.exporter import export_exceptions
from utils.validator import validate_outputs


def _verify_boundaries(file_path: Path, logger: logging.Logger) -> bool:
    logger.info("Executing boundary verification matrix scan...")
    wb_ro = openpyxl.load_workbook(file_path, read_only=True)

    for sheet_name in wb_ro.sheetnames:
        ws_ro = wb_ro[sheet_name]

        row_count = 0
        for _ in ws_ro.iter_rows():
            row_count += 1
            if row_count > MAX_ROWS:
                logger.error(f"Row limit breached in {sheet_name}. Terminating to prevent OOM.")
                return False

        if ws_ro.max_column and ws_ro.max_column > MAX_COLS:
            logger.error(f"Column limit breached in {sheet_name}. Terminating to prevent OOM.")
            return False

    wb_ro.close()
    return True


def process_workbook(file_path: Path, logger: logging.Logger) -> bool:
    if not _verify_boundaries(file_path, logger):
        return False

    logger.info("Ingesting target workbook...")
    wb = openpyxl.load_workbook(file_path)

    # Load identical state to ensure baseline maps formulas strictly as strings
    wb_master = openpyxl.load_workbook(file_path)

    exceptions: list[dict[str, str]] = []
    master_sheet_names: list[str] = wb.sheetnames
    total_panels_created: int = 0
    global_all_panel_cols: set[int] = set()

    for sheet_name in master_sheet_names:
        ws = wb[sheet_name]
        logger.info(f"Targeting matrix: {sheet_name}")

        panel_cols, all_panel_cols, header_map = detect_columns(ws, logger)
        global_all_panel_cols.update(all_panel_cols)

        if not panel_cols:
            logger.warning(f"No series detected in {sheet_name}.")
            exceptions.append({
                "sheet": sheet_name,
                "issue": "Series Void",
                "description": "Row boundary scan returned no valid panel nomenclature."
            })
            continue

        count = classify_panels(wb, ws, panel_cols, all_panel_cols, header_map, logger)
        total_panels_created += count

        del wb[sheet_name]

    if total_panels_created == 0:
        logger.error("Architecture operation failed. Zero distinct panels generated.")
        return False

    out_path: Path = OUTPUT_DIR / OUTPUT_FILE
    wb.save(out_path)
    logger.info(f"Workbook compiled and locked at {out_path}")

    validate_outputs(wb_master, wb, logger)
    export_exceptions(exceptions, logger)

    return True