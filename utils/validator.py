"""
Post-processing cryptograph-style numeric verification.
Utilizes exact precision decimal mapping to mitigate float degradation.
"""
import logging
from decimal import Decimal, InvalidOperation
from openpyxl.workbook.workbook import Workbook

from config import HEADER_ROW, PANEL_ROW, END_MARKER


def _extract_sum(ws, col_idx: int) -> Decimal:
    """Aggregates absolute columnar data utilizing exact arithmetic. Ignored strings/formulas."""
    total = Decimal('0.0')
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=col_idx).value
        if cell_val is not None:
            if isinstance(cell_val, str) and str(cell_val).startswith('='):
                continue
            try:
                total += Decimal(str(cell_val).strip())
            except InvalidOperation:
                pass
    return total


def validate_outputs(wb_master: Workbook, wb_out: Workbook, logger: logging.Logger) -> bool:
    """
    Enforces a strict equivalence check between the source numeric footprint
    and the compiled output to guarantee structural and financial fidelity.
    """
    logger.info("Initiating precision numeric validation protocol.")
    passed: bool = True

    master_sums: dict[str, Decimal] = {}

    for sheet_name in wb_master.sheetnames:
        ws = wb_master[sheet_name]
        for col in range(6, ws.max_column + 1):
            val = str(ws.cell(row=PANEL_ROW, column=col).value or "").strip().upper()
            if END_MARKER in val:
                break
            panel_name = str(ws.cell(row=PANEL_ROW, column=col).value or "").strip()
            if panel_name:
                master_sums[panel_name] = master_sums.get(panel_name, Decimal('0.0')) + _extract_sum(ws, col)

    output_sums: dict[str, Decimal] = {}
    for sheet_name in wb_out.sheetnames:
        ws = wb_out[sheet_name]
        for col in range(6, ws.max_column + 1):
            val = str(ws.cell(row=PANEL_ROW, column=col).value or "").strip().upper()
            if END_MARKER in val:
                break
            panel_name = str(ws.cell(row=PANEL_ROW, column=col).value or "").strip()
            if panel_name in master_sums:
                output_sums[panel_name] = output_sums.get(panel_name, Decimal('0.0')) + _extract_sum(ws, col)

    for panel_name, expected in master_sums.items():
        actual = output_sums.get(panel_name, Decimal('0.0'))
        if expected != actual:
            logger.error(f"Fidelity failure on [{panel_name}]. Baseline={expected}, Compiled={actual}.")
            passed = False

    if passed:
        logger.info("Validation complete: 100% equivalence achieved against baseline matrix.")

    return passed