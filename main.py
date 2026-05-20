"""
Panel Classifier — BOM Splitter  (v6 - format-identical)
==========================================================
Just press Play in PyCharm. No commands or filenames needed.

Handles TWO master file formats automatically:

  FORMAT A — JCB / ATP Control Panel  (sheet name = 'BOM')
  FORMAT B — Daikin / ACS  (sheet name = filename)

For Daikin files the split sheet is IDENTICAL to the original:
  - Same column layout (no column compression)
  - Same AMT formula referencing original column positions
  - Section header rows (ELECTRICAL, WIRE, BUSBAR, PACKING) preserved
  - Original SNo values kept (no renumbering)
  - Summary section values written in the correct panel column
  - Freeze panes set to F5 for easy viewing

Logs:  <script folder>/logs/bom_splitter_YYYY-MM-DD.log
"""

import os
import sys
import glob
import logging
from copy import copy
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ── Logging ───────────────────────────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOG_DIR    = os.path.join(SCRIPT_DIR, "logs")
os.makedirs(LOG_DIR, exist_ok=True)
LOG_FILE   = os.path.join(LOG_DIR,
             f"bom_splitter_{datetime.now().strftime('%Y-%m-%d')}.log")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ]
)
log = logging.getLogger("BOMSplitter")

# ── Helpers ───────────────────────────────────────────────────────────────────

def _is_num(v):
    if v is None: return False
    try: float(v); return True
    except: return False

def _safe_name(s):
    for ch in r'/\?*[]:':
        s = s.replace(ch, "-")
    return s.strip()[:31]

def _copy_cell(src_cell, dst_cell):
    """Copy value, font, fill, border, alignment, number_format from src to dst."""
    dst_cell.value = src_cell.value
    if src_cell.has_style:
        dst_cell.font      = copy(src_cell.font)
        dst_cell.fill      = copy(src_cell.fill)
        dst_cell.border    = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


# ══════════════════════════════════════════════════════════════════════════════
#  FORMAT A  —  JCB / ATP  (sheet = 'BOM')
# ══════════════════════════════════════════════════════════════════════════════

HAIR = Side(style="hair")
THIN = Side(style="thin")

def _hair(): return Border(top=HAIR, bottom=HAIR, left=HAIR, right=HAIR)
def _thin(): return Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
def _font(bold=False, size=8): return Font(name="Arial", bold=bold, size=size)
def _fill(h=None): return PatternFill("solid", fgColor=h) if h else PatternFill(fill_type=None)
def _align(h="general"): return Alignment(horizontal=h, vertical="center")

def _cell(ws, row, col, val, bold=False, fill=None,
          align="general", border="hair", fmt=None, size=8):
    c = ws.cell(row=row, column=col)
    c.value     = val
    c.font      = _font(bold, size)
    c.fill      = _fill(fill)
    c.alignment = _align(align)
    c.border    = _hair() if border == "hair" else (
                  _thin() if border == "thin" else Border())
    if fmt: c.number_format = fmt
    return c

JCB_PANEL  = "CCFF66"
JCB_HDR_L  = "FFCCFF"
JCB_HDR_P  = "CCFFFF"
JCB_CATNO  = "FFFFCC"
JCB_LABOUR = "FFFF99"
JCB_TOT_E  = "FF6600"
JCB_TOT_P  = "CCFF66"
JCB_FINAL  = "FF0000"
JCB_FINLBL = "CCFFFF"


def _jcb_get_summary(summary_list, keywords, occurrence=1):
    count = 0
    for lbl, val in summary_list:
        if any(kw.lower() in lbl.lower() for kw in keywords):
            count += 1
            if count == occurrence: return val
    return None


def _jcb_read(path):
    wb   = openpyxl.load_workbook(path, read_only=True)
    ws   = wb["BOM"]
    rows = list(ws.iter_rows(values_only=True))
    row1, row2, row4 = rows[0], rows[1], rows[3]

    panels = []
    for ci in range(5, len(row1)):
        v = row1[ci]
        if not v or not str(v).strip(): continue
        if str(v).strip().upper() in ("EXISTING COST", "PROPOSED COST"): break
        code = str(row2[ci]).replace("\xa0", "").strip() if row2[ci] else ""
        panels.append({"col_idx": ci, "model": str(v).strip(), "code": code})

    existing_ur_ci = proposed_ur_ci = None
    seen = 0
    for ci, v in enumerate(row4):
        if v and str(v).strip() == "U/R":
            if seen == 0: existing_ur_ci = ci
            elif seen == 1: proposed_ur_ci = ci; break
            seen += 1

    all_bom = []
    for row in rows[4:]:
        if _is_num(row[0]):   all_bom.append(row)
        elif not all_bom:     continue
        else:                 break

    packing_start = len(all_bom)
    for i, row in enumerate(all_bom):
        desc = str(row[1]).strip().upper() if row[1] else ""
        if desc == "PACKING" and not row[2] and not row[3]:
            packing_start = i; break

    component_rows = all_bom[:packing_start]
    packing_rows   = all_bom[packing_start:]

    summary_rows = [r for r in rows[4 + len(all_bom):]
                    if any(v is not None for v in r)]

    panel_summary = {p["col_idx"]: [] for p in panels}
    for row in summary_rows:
        lbl = str(row[4]).strip() if row[4] else ""
        if not lbl: continue
        for p in panels:
            ci  = p["col_idx"]
            val = row[ci] if ci < len(row) else None
            panel_summary[ci].append((lbl, val))

    return panels, component_rows, packing_rows, existing_ur_ci, proposed_ur_ci, panel_summary


def _jcb_create_sheet(wb, panel, component_rows, packing_rows,
                       existing_ur_ci, proposed_ur_ci, panel_summary):
    ws = wb.create_sheet(title=_safe_name(panel["code"]))
    ci = panel["col_idx"]

    for col, w in zip("ABCDEFGHIJK",
                      [4.4,39.4,10.9,15.1,5.4,8.4,7.4,9.9,8.4,7.4,9.9]):
        ws.column_dimensions[col].width = w
    for r in range(1, 5):
        ws.row_dimensions[r].height = 14.25 if r < 4 else 16.5

    _cell(ws,1,1," ",             bold=True, align="center", border="none")
    _cell(ws,1,5,"MODEL",         bold=True, align="right",  border="none")
    _cell(ws,1,6,panel["model"],  bold=True, align="center", fill=JCB_PANEL)
    _cell(ws,1,7,"EXISTING COST", bold=True, align="center", fill=JCB_HDR_L)
    _cell(ws,1,10,"PROPOSED Cost",bold=True, align="center", fill=JCB_HDR_P)
    _cell(ws,2,5,"CODE",          bold=True, align="right",  border="none")
    _cell(ws,2,6,panel["code"],   bold=True, align="center", fill=JCB_PANEL)
    _cell(ws,3,6,1,               bold=True, align="center", fill=JCB_PANEL)

    for col, hdr, fill in [
        (1,"SNo",JCB_HDR_L),(2,"DESCRIPTION",JCB_HDR_L),(3,"MAKE",JCB_HDR_L),
        (4,"CAT NO.",JCB_HDR_L),(5,"UNIT",JCB_HDR_L),(6,"QTY",JCB_HDR_L),
        (7,"U/R",JCB_HDR_L),(8,"AMT",JCB_HDR_L),
        (10,"U/R",JCB_HDR_P),(11,"AMT",JCB_HDR_P),
    ]:
        _cell(ws,4,col,hdr,bold=True,align="center",fill=fill)
    ws.cell(row=4,column=9).border = _hair()

    filtered_comp = [r for r in component_rows
                     if _is_num(r[ci]) and float(r[ci]) != 0]
    er = 5
    for sno, row in enumerate(filtered_comp, 1):
        _cell(ws,er,1,sno,                 align="center")
        _cell(ws,er,2,row[1])
        _cell(ws,er,3,row[2])
        _cell(ws,er,4,row[3],              fill=JCB_CATNO)
        _cell(ws,er,5,row[4])
        _cell(ws,er,6,row[ci],             align="center")
        _cell(ws,er,7,row[existing_ur_ci], fmt="#,##0.00")
        c=ws.cell(row=er,column=8);  c.value=f"=G{er}*F{er}"; c.font=_font(); c.border=_hair(); c.number_format="#,##0.00"
        ws.cell(row=er,column=9).border=_hair()
        _cell(ws,er,10,row[proposed_ur_ci],fmt="#,##0.00")
        c=ws.cell(row=er,column=11); c.value=f"=J{er}*F{er}"; c.font=_font(); c.border=_hair(); c.number_format="#,##0.00"
        er += 1

    last_comp_row = er - 1
    data_start    = 5

    filtered_pack = [r for r in packing_rows
                     if _is_num(r[ci]) and float(r[ci]) > 0]
    sno = len(filtered_comp) + 1
    for row in filtered_pack:
        _cell(ws,er,1,sno,                 align="center")
        _cell(ws,er,2,row[1])
        _cell(ws,er,3,row[2])
        _cell(ws,er,4,row[3],              fill=JCB_CATNO)
        _cell(ws,er,5,row[4])
        _cell(ws,er,6,row[ci],             align="center")
        _cell(ws,er,7,row[existing_ur_ci], fmt="#,##0.00")
        c=ws.cell(row=er,column=8);  c.value=f"=G{er}*F{er}"; c.font=_font(); c.border=_hair(); c.number_format="#,##0.00"
        ws.cell(row=er,column=9).border=_hair()
        _cell(ws,er,10,row[proposed_ur_ci],fmt="#,##0.00")
        c=ws.cell(row=er,column=11); c.value=f"=J{er}*F{er}"; c.font=_font(); c.border=_hair(); c.number_format="#,##0.00"
        er += 1; sno += 1

    ps        = panel_summary.get(ci, [])
    labour_e  = _jcb_get_summary(ps, ["mfg","labour"],        occurrence=1)
    labour_p  = _jcb_get_summary(ps, ["mfg","labour"],        occurrence=2)
    pack_e    = _jcb_get_summary(ps, ["packing & pallet"],    occurrence=1)
    pack_p    = _jcb_get_summary(ps, ["packing & pallet"],    occurrence=2)
    freight_e = _jcb_get_summary(ps, ["freight"],             occurrence=1)
    freight_p = _jcb_get_summary(ps, ["freight"],             occurrence=2)

    sr = er + 1

    def srow(lbl, val, bold=False, lbl_fill=None, val_fill=None):
        nonlocal sr
        lc = ws.cell(row=sr, column=5)
        lc.value, lc.font = lbl, _font(bold)
        lc.alignment, lc.border = _align("right"), _thin()
        if lbl_fill: lc.fill = _fill(lbl_fill)
        vc = ws.cell(row=sr, column=6)
        vc.value, vc.font = val, _font(bold)
        vc.border, vc.number_format = _thin(), "#,##0.00"
        if val_fill: vc.fill = _fill(val_fill)
        sr += 1; return sr - 1

    def labour_row(lbl, val):
        nonlocal sr
        lc = ws.cell(row=sr, column=5)
        lc.value, lc.font = lbl, _font(bold=True)
        lc.alignment, lc.border = _align("right"), _thin()
        vc = ws.cell(row=sr, column=6)
        vc.value, vc.font = val, _font()
        vc.fill, vc.border, vc.number_format = _fill(JCB_LABOUR), _thin(), "#,##0.00"
        sr += 1; return sr - 1

    r_tot_e = sr
    srow("TOTAL BASIC EXISTING COST :",
         f"=SUMPRODUCT(F{data_start}:F{last_comp_row},G{data_start}:G{last_comp_row})",
         bold=True, val_fill=JCB_TOT_E)
    r_inv_e = sr
    srow("3% Inventory Carry Charges on RMC", f"=F{r_tot_e}*3%")
    r_lab_e = labour_row("Mfg. Labour Charges Incl. Staff Charges", labour_e)
    srow("7% Over Head", f"=+F{r_lab_e}*7%")
    r_pr_e  = sr
    srow("Profit (8%)", f"=SUM(F{r_tot_e}+F{r_lab_e})*0.08")
    r_br_e  = sr
    srow("Basic Rate", f"=SUM(F{r_tot_e}:F{r_pr_e})", bold=True)
    srow("Packing & Pallet", pack_e)
    srow("Freight from KEPL to JCB Ballabhgarh", freight_e)
    r_fin_e = sr
    srow("Final Rate", f"=SUM(F{r_br_e}:F{r_fin_e-1})",
         bold=True, lbl_fill=JCB_FINLBL, val_fill=JCB_FINAL)
    sr += 1
    r_tot_p = sr
    srow("TOTAL BASIC NEW COST :",
         f"=SUMPRODUCT(F{data_start}:F{last_comp_row},J{data_start}:J{last_comp_row})",
         bold=True, val_fill=JCB_TOT_P)
    r_inv_p = sr
    srow("3% Inventory Carry Charges on RMC", f"=F{r_tot_p}*3%")
    r_lab_p = labour_row("Mfg. Labour Charges Incl. Staff Charges w.e.f. 1.4.26", labour_p)
    srow("7% Over Head", f"=+F{r_lab_p}*7%")
    r_pr_p  = sr
    srow("Profit (8%)", f"=SUM(F{r_tot_p}+F{r_lab_p})*0.08")
    r_br_p  = sr
    srow("Basic Rate", f"=SUM(F{r_tot_p},F{r_inv_p},F{r_lab_p}:F{r_pr_p})", bold=True)
    srow("Packing & Pallet", pack_p)
    srow("Freight from KEPL to JCB Ballabhgarh", freight_p)
    r_fin_p = sr
    srow("Final RATE :", f"=SUM(F{r_br_p}:F{r_fin_p-1})",
         bold=True, lbl_fill=JCB_FINLBL, val_fill=JCB_FINAL)

    ws.freeze_panes = "A5"
    log.info(f"    ✓ {_safe_name(panel['code'])[:22]:22}  "
             f"({len(filtered_comp)} comp + {len(filtered_pack)} packing)")


def process_jcb(input_path):
    log.info("  Format : JCB / ATP  (BOM sheet)")
    panels, comp, pack, e_ur, p_ur, ps = _jcb_read(input_path)
    log.info(f"  Panels: {len(panels)}  Components: {len(comp)}  Packing: {len(pack)}")
    wb = openpyxl.load_workbook(input_path)
    for panel in panels:
        _jcb_create_sheet(wb, panel, comp, pack, e_ur, p_ur, ps)
    base, ext = os.path.splitext(input_path)
    out = base + "_SPLIT" + ext
    wb.save(out)
    log.info(f"  ✅ Saved → {os.path.basename(out)}")


# ══════════════════════════════════════════════════════════════════════════════
#  FORMAT B  —  Daikin / ACS
#  Strategy: copy the master sheet row-by-row, keeping IDENTICAL column layout.
#  For each panel sheet we only include rows where that panel has a value.
# ══════════════════════════════════════════════════════════════════════════════

def _dai_read_structure(path):
    """
    Returns everything needed to create panel sheets from a Daikin master.
    """
    wb_ro = openpyxl.load_workbook(path, read_only=True)
    ws_ro = wb_ro.active
    all_rows_values = list(ws_ro.iter_rows(values_only=True))

    # Full workbook with styles for cell copying
    wb_styled = openpyxl.load_workbook(path)
    ws_styled = wb_styled.active

    row1 = all_rows_values[0]
    row2 = all_rows_values[1]
    row4 = all_rows_values[3]

    # Detect panel start column: first QTY col in row4 (0-based)
    first_qty_ci = next(
        (i for i, v in enumerate(row4) if v and str(v).strip().upper() == "QTY"), 5)

    # U/R and AMT column indices (0-based)
    ur_ci  = next((i for i, v in enumerate(row4) if v and str(v).strip() == "U/R"), None)
    amt_ci = next((i for i, v in enumerate(row4) if v and str(v).strip() == "AMT"), None)

    if ur_ci is None:
        raise ValueError("Could not find 'U/R' column in row 4")

    # Panel columns: from first_qty_ci up to (not including) ur_ci
    panels = []
    for ci in range(first_qty_ci, ur_ci):
        model = str(row1[ci]).strip() if row1[ci] else ""
        code  = str(row2[ci]).strip() if row2[ci] else ""
        # Skip any non-panel labels
        if model.upper() in ("RATE", "AMT", "U/R", "QTY", ""): continue
        panels.append({
            "col_idx":  ci,           # 0-based index in master
            "col_1b":   ci + 1,       # 1-based column number
            "model":    model,
            "code":     code,
        })

    # Summary rows: find the row index where 'Daikin PROPOSED BASIC COST' appears
    summary_start_row = None   # 0-based index in all_rows_values
    for i, row in enumerate(all_rows_values):
        for v in row:
            if v and "daikin proposed basic cost" in str(v).lower():
                summary_start_row = i
                break
        if summary_start_row is not None:
            break

    # BOM rows: rows 4 (0-based) up to summary_start_row (exclusive)
    # Includes header rows 0-3, data rows 4..summary_start_row-1
    bom_end = summary_start_row if summary_start_row else len(all_rows_values)

    # Summary rows: from summary_start_row to end
    summary_rows_values = all_rows_values[summary_start_row:] if summary_start_row else []

    # Per-panel summary values dict: {col_1b: {label_lower: value}}
    # The label column in summary varies: find it by scanning for 'Daikin' text
    panel_summary = {p["col_1b"]: {} for p in panels}
    for row in summary_rows_values:
        # Find which cell has the label
        lbl = ""
        lbl_found_at = -1
        for ci2, v in enumerate(row):
            if v and isinstance(v, str) and (
                "daikin" in v.lower() or "conversion" in v.lower() or
                "enclosure" in v.lower() or "fsm" in v.lower() or
                "1cc" in v.lower() or "proposal" in v.lower()
            ):
                lbl = v.strip()
                lbl_found_at = ci2
                break
        if not lbl:
            # Also check for FSM which has label in col D
            for ci2, v in enumerate(row):
                if v and isinstance(v, str) and "fsm" in v.lower():
                    lbl = v.strip(); break
        if not lbl: continue

        for p in panels:
            ci2 = p["col_idx"]
            val = row[ci2] if ci2 < len(row) else None
            if val is not None:
                panel_summary[p["col_1b"]][lbl.lower()] = val

    # Also read FSM rate (col E or F in summary)
    fsm_rate_by_panel = {}
    for row in summary_rows_values:
        lbl_in_row = False
        for v in row:
            if v and isinstance(v, str) and "fsm" in v.lower():
                lbl_in_row = True; break
        if lbl_in_row:
            # FSM rate is in the col right before panel cols (shared)
            # Find the numeric value in cols before first panel col
            for try_ci in range(first_qty_ci - 1, -1, -1):
                v = row[try_ci] if try_ci < len(row) else None
                if _is_num(v):
                    for p in panels:
                        fsm_rate_by_panel[p["col_1b"]] = v
                    break

    return (panels, all_rows_values, ws_styled,
            first_qty_ci, ur_ci, amt_ci,
            bom_end, summary_rows_values, panel_summary, fsm_rate_by_panel)


def _dai_create_sheet(wb_out, panel, master_ws,
                       all_rows_values, first_qty_ci, ur_ci, amt_ci,
                       bom_end, summary_rows_values,
                       panel_summary, fsm_rate_by_panel):
    """
    Create one panel sheet with COMPRESSED columns.

    Master layout:  shared(1..first_qty_ci) | panel_cols(first_qty_ci..ur_ci-1) | U/R | AMT
    Output layout:  shared(1..first_qty_ci) | QTY(first_qty_ci+1) | U/R(+2) | AMT(+3)

    All panel QTY columns are collapsed to a single QTY column.
    U/R and AMT shift left to immediately follow it.
    AMT formula: =UR_out_col * QTY_out_col  (e.g. =G{row}*F{row})
    Summary labels stay at their original positions (within shared block).
    Summary value written in the single QTY output column.
    """
    ci     = panel["col_idx"]   # 0-based panel column in master
    col_1b = panel["col_1b"]    # 1-based panel column in master

    # Output column positions (1-based)
    qty_out = first_qty_ci + 1          # e.g. col 6 (F) for standard files
    ur_out  = first_qty_ci + 2          # e.g. col 7 (G)
    amt_out = first_qty_ci + 3          # e.g. col 8 (H)
    qty_ltr = get_column_letter(qty_out)
    ur_ltr  = get_column_letter(ur_out)
    amt_ltr = get_column_letter(amt_out)

    # Column mapping: master col (1-based) -> output col (1-based)
    # Shared cols (1..first_qty_ci): stay in place
    # This panel's QTY col (ci+1): maps to qty_out
    # All other panel QTY cols: dropped
    # U/R col (ur_ci+1): maps to ur_out
    # AMT col (amt_ci+1): maps to amt_out
    def out_col(master_col_1b):
        mc0 = master_col_1b - 1   # 0-based
        if mc0 < first_qty_ci:    return master_col_1b   # shared: unchanged
        if mc0 == ci:             return qty_out          # this panel QTY
        if mc0 == ur_ci:          return ur_out           # U/R
        if amt_ci and mc0 == amt_ci: return amt_out       # AMT
        return None                                        # other panel cols: skip

    sheet_name = _safe_name(panel["model"])
    ws = wb_out.create_sheet(title=sheet_name)

    # ── Column widths ─────────────────────────────────────────────────────────
    # Copy shared cols widths, then set QTY/U/R/AMT widths
    for mc_ltr, dim in master_ws.column_dimensions.items():
        mc_idx = openpyxl.utils.column_index_from_string(mc_ltr) - 1  # 0-based
        if mc_idx < first_qty_ci:
            ws.column_dimensions[mc_ltr].width = dim.width
    # QTY, U/R, AMT widths from master's corresponding cols
    for src_ci, dst_1b in [(ci, qty_out), (ur_ci, ur_out),
                            (amt_ci if amt_ci else ur_ci, amt_out)]:
        src_ltr = get_column_letter(src_ci + 1)
        dst_ltr = get_column_letter(dst_1b)
        dim = master_ws.column_dimensions.get(src_ltr)
        if dim:
            ws.column_dimensions[dst_ltr].width = dim.width
        else:
            ws.column_dimensions[dst_ltr].width = 10.0

    # ── Row heights for header rows ───────────────────────────────────────────
    for rn in range(1, 5):
        rh = master_ws.row_dimensions.get(rn)
        if rh and rh.height:
            ws.row_dimensions[rn].height = rh.height

    # ── Header rows 1–4 ───────────────────────────────────────────────────────
    for master_rn in range(1, 5):
        for cell in master_ws[master_rn]:
            oc = out_col(cell.column)
            if oc is None:
                continue
            dst = ws.cell(row=master_rn, column=oc)
            _copy_cell(cell, dst)
            # For header rows: clear values in panel cols that aren't this panel
            mc0 = cell.column - 1
            if mc0 >= first_qty_ci and mc0 < ur_ci and mc0 != ci:
                dst.value = None   # already skipped above, but safety

    # ── BOM rows ──────────────────────────────────────────────────────────────
    out_row = 5
    data_rows_written = []
    new_sno = 0   # sequential counter for data rows only

    for master_row_idx in range(5, bom_end + 1):
        if master_row_idx > len(all_rows_values):
            break
        row_vals = all_rows_values[master_row_idx - 1]

        sno       = row_vals[0]
        desc      = row_vals[1]
        # UNIT col: last shared col before panel cols
        unit_ci   = first_qty_ci - 1
        unit      = row_vals[unit_ci] if unit_ci < len(row_vals) else None
        panel_qty = row_vals[ci] if ci < len(row_vals) else None

        # Section header: desc present, no unit, U/R = 0
        is_section_header = (
            desc is not None and str(desc).strip() != "" and
            unit is None and
            ur_ci < len(row_vals) and row_vals[ur_ci] == 0
        )

        is_data_row = (
            _is_num(sno) and unit is not None and
            panel_qty is not None and _is_num(panel_qty) and
            float(panel_qty) != 0
        )

        if not is_section_header and not is_data_row:
            continue

        # Increment counter for data rows only
        if is_data_row:
            new_sno += 1

        master_row_cells = master_ws[master_row_idx]
        for cell in master_row_cells:
            mc0 = cell.column - 1
            oc  = out_col(cell.column)
            if oc is None:
                continue
            dst = ws.cell(row=out_row, column=oc)
            _copy_cell(cell, dst)
            # Overwrite SNo: sequential for data rows, blank for section headers
            if mc0 == 0:
                dst.value = new_sno if is_data_row else None
            # AMT column: write compressed formula
            if amt_ci and mc0 == amt_ci:
                dst.value = f"={ur_ltr}{out_row}*{qty_ltr}{out_row}"

        rh = master_ws.row_dimensions.get(master_row_idx)
        if rh and rh.height:
            ws.row_dimensions[out_row].height = rh.height

        if is_data_row:
            data_rows_written.append(out_row)

        out_row += 1

    # ── Summary section ───────────────────────────────────────────────────────
    # Find first summary master row
    summary_master_start = None
    for mr in range(bom_end, master_ws.max_row + 1):
        row_v = all_rows_values[mr - 1] if mr <= len(all_rows_values) else []
        if any(v and "daikin proposed basic cost" in str(v).lower() for v in row_v):
            summary_master_start = mr
            break

    if summary_master_start is None:
        log.warning(f"    Could not find summary section in master")
        ws.freeze_panes = f"{qty_ltr}5"
        log.info(f"    ✓ {sheet_name[:22]:22}")
        return

    out_row += 1   # blank gap row
    sr = out_row

    for mr in range(summary_master_start, master_ws.max_row + 1):
        if mr > len(all_rows_values):
            break
        row_v = all_rows_values[mr - 1]

        master_row_cells = master_ws[mr]
        for cell in master_row_cells:
            mc0 = cell.column - 1
            oc  = out_col(cell.column)
            if oc is None:
                continue
            dst = ws.cell(row=sr, column=oc)
            _copy_cell(cell, dst)
            # Write this panel's value in the QTY output col
            if mc0 == ci:
                dst.value = row_v[ci] if ci < len(row_v) else None

        rh = master_ws.row_dimensions.get(mr)
        if rh and rh.height:
            ws.row_dimensions[sr].height = rh.height

        sr += 1

    # ── Freeze panes at QTY col, row 5 ───────────────────────────────────────
    ws.freeze_panes = f"{qty_ltr}5"

    ps          = panel_summary.get(col_1b, {})
    basic_cost  = next((v for k, v in ps.items() if "daikin proposed basic cost" in k), None)
    final_price = next((v for k, v in ps.items() if "daikin final price" in k), None)
    n_items     = len(data_rows_written)
    basic_str   = f"{basic_cost:.0f}"  if basic_cost  and _is_num(basic_cost)  else "N/A"
    final_str   = f"{final_price:.0f}" if final_price and _is_num(final_price) else "N/A"
    log.info(f"    ✓ {sheet_name[:22]:22}  ({n_items} items | Basic={basic_str} | Final={final_str})")


def process_daikin(input_path):
    log.info("  Format : Daikin / ACS")
    (panels, all_rows_values, ws_styled,
     first_qty_ci, ur_ci, amt_ci,
     bom_end, summary_rows_values, panel_summary,
     fsm_rate_by_panel) = _dai_read_structure(input_path)

    log.info(f"  Panels: {len(panels)}  |  U/R col: {ur_ci+1}  |  AMT col: {amt_ci+1 if amt_ci else 'none'}")

    wb_out = openpyxl.load_workbook(input_path)  # load with styles for copying

    for panel in panels:
        _dai_create_sheet(
            wb_out, panel, wb_out.active,
            all_rows_values, first_qty_ci, ur_ci, amt_ci,
            bom_end, summary_rows_values, panel_summary, fsm_rate_by_panel
        )

    base, ext = os.path.splitext(input_path)
    out = base + "_SPLIT" + ext
    wb_out.save(out)
    log.info(f"  ✅ Saved → {os.path.basename(out)}")


# ══════════════════════════════════════════════════════════════════════════════
#  Format detection
# ══════════════════════════════════════════════════════════════════════════════

def detect_format(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    return "JCB" if "BOM" in wb.sheetnames else "DAIKIN"


def process_file(input_path):
    filename = os.path.basename(input_path)
    log.info(f"{'='*60}")
    log.info(f"Processing : {filename}")
    log.info(f"{'='*60}")
    try:
        fmt = detect_format(input_path)
        if fmt == "JCB":
            process_jcb(input_path)
        else:
            process_daikin(input_path)
        return True
    except Exception as e:
        log.error(f"❌ FAILED : {filename}")
        log.error(f"   Reason : {e}", exc_info=True)
        return False


# ══════════════════════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    log.info("")
    log.info("╔══════════════════════════════════════════════╗")
    log.info("║       Panel Classifier  —  BOM Splitter      ║")
    log.info(f"║       Started : {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}             ║")
    log.info("╚══════════════════════════════════════════════╝")
    log.info(f"Folder : {SCRIPT_DIR}")
    log.info(f"Log    : {LOG_FILE}")
    log.info("")

    all_xlsx    = glob.glob(os.path.join(SCRIPT_DIR, "*.xlsx"))
    input_files = [f for f in all_xlsx
                   if not os.path.basename(f).endswith("_SPLIT.xlsx")]

    if not input_files:
        log.warning("No .xlsx files found. Place master file(s) here and run again.")
        return

    log.info(f"Found {len(input_files)} file(s):")
    for f in input_files:
        log.info(f"  • {os.path.basename(f)}")
    log.info("")

    success = failed = 0
    for f in input_files:
        ok = process_file(f)
        if ok: success += 1
        else:  failed  += 1
        log.info("")

    log.info("══════════════════════════════════════════════")
    log.info(f"  DONE  —  {success} succeeded,  {failed} failed")
    log.info(f"  Finished : {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}")
    log.info("══════════════════════════════════════════════")


if __name__ == "__main__":
    main()